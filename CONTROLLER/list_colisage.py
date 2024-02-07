import customtkinter as ctk
import sys
import os
import getpass

# Obtenez le nom d'utilisateur actuel
current_user = getpass.getuser()

from openpyxl import load_workbook
from MODEL import DatabaseSetup as db
from PIL import ImageTk, Image
# Fonction pour enregistrer les données dans le fichier Excel
# Import StringVar from tkinter
from CTkMessagebox import CTkMessagebox
from CONTROLLER import cLogin as cl
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
def enregistrer(frame, champ_nom_expediteur, champ_nom_entreprise, champ_adress_entre, champ_Tel_entre,
               champ_email_entre, champ_nom_destinataire, champ_Contact, champ_adress_destinataire,
               champ_Tel_destinataire, champ_portabl_destinataire, champ_email_destinataire,
               champ_numero_facture, champ_numero_awb, champ_nombre_colis, champ_raison, champ_details,fenetre,root):
    nombre_colis = int(champ_nombre_colis)
    if nombre_colis <= 0:
        label = ctk.CTkLabel(master=frame, text="Veuillez entrer un nombre valide", text_color="red", font=('bold', 10))
        label.grid(row=15, column=2)
    else:
        for widget in frame.winfo_children():
            widget.destroy()
        parcel_data = []
        print("nombre_colis "+str(nombre_colis))
        label = ctk.CTkLabel(frame, text="Description")
        label.grid(row=1, column=0, padx=5, pady=5)
        label = ctk.CTkLabel(frame, text="le nombre des produits")
        label.grid(row=3, column=0, padx=5, pady=5)
        label = ctk.CTkLabel(frame, text="les dimensions")
        label.grid(row=5, column=0, padx=5, pady=5)
        label = ctk.CTkLabel(frame, text="les marques ")
        label.grid(row=7, column=0, padx=5, pady=5)
        
        for i in range(nombre_colis):
            # Create a dictionary to store data for each parcel
            parcel = {
              "Description":None,
              "nombre_produits": None,
              "dimensions": None,
              "marque_colis": None
            }

            # Collect user input for each parcel
            parcel["Description"] = ctk.CTkEntry(frame, placeholder_text=f" pour colis numero {i}:")
            parcel["nombre_produits"] = ctk.CTkEntry(frame, placeholder_text=f" pour colis numero {i}:")
            parcel["dimensions"] = ctk.CTkEntry(frame,placeholder_text=f"pour colis numero {i}:")
            parcel["marque_colis"] = ctk.CTkEntry(frame,placeholder_text=f" pour colis numero {i}:")

            # Use grid to place the entry widgets
            parcel["Description"].grid(row=2, column=i, padx=5, pady=5)
            parcel["nombre_produits"].grid(row=4, column=i, padx=5, pady=5)
            parcel["dimensions"].grid(row=6, column=i, padx=5, pady=5)
            parcel["marque_colis"].grid(row=8, column=i, padx=5, pady=5)

            # Add this parcel's data to the list
            print(f"parcel {parcel}")
            parcel_data.append(parcel)

        print(f"parcel_data {parcel_data}")
        # Create a button to calculate total weight
        bouton_soumettre = ctk.CTkButton(frame, text="Entrer", command=lambda:  create_product_reference_comboboxes(frame, nombre_colis,parcel_data))
        bouton_soumettre.grid()
        bouton_enregistrer = ctk.CTkButton(frame, text="Retourner", command=lambda:cl.retourner_Login(root))
        bouton_enregistrer.grid(pady=5)
       

        # Charger le modèle de liste de colisage Excel
        wb = load_workbook(resource_path('CONTROLLER\\Modèle-liste-de-colisage-français.xlsx'))
        ws = wb.active
    
        # Récupérer les données saisies par l'utilisateur depuis les variables tkinter
        nom_expediteur = champ_nom_expediteur
        nom_entreprise = champ_nom_entreprise
        adress_entreprise = champ_adress_entre
        Tel_entreprise = champ_Tel_entre
        email_entreprise = champ_email_entre
        nom_destinataire = champ_nom_destinataire
        contact_destinataire = champ_Contact
        adress_destinataire = champ_adress_destinataire
        Tel_destinataire = champ_Tel_destinataire
        portabl_destinataire = champ_portabl_destinataire
        email_destinataire = champ_email_destinataire
        #date = champ_date.get()
        numero_facture = champ_numero_facture
        numero_awb = champ_numero_awb
        #poids_brut= champ_poids_brut.get()
        nombre_total_colis= champ_nombre_colis
        raison_export = champ_raison
        autre_details = champ_details
    
    
        # Écrire les données dans les cellules appropriées du modèle Excel
        ws['D7'] = nom_expediteur
        ws['D8'] = nom_entreprise
        ws['D9'] = adress_entreprise
        ws['D11'] = Tel_entreprise
        ws['D12'] = email_entreprise
        ws['D15'] = nom_destinataire
        ws['D16'] = contact_destinataire
        ws['D17'] = adress_destinataire
        ws['D19'] = Tel_destinataire
        ws['D20'] = portabl_destinataire
        ws['D21'] = email_destinataire
        ws['F8'] = numero_facture
        #ws['F4'] = date
        ws['F10'] = numero_awb
        #ws['F12'] = poids_brut
        ws['F14'] = nombre_total_colis
        ws['F16'] = raison_export
        ws['F18'] = autre_details
    
    

        # Enregistrer le fichier Excel
        wb.save(resource_path('CONTROLLER\\Modèle-liste-de-colisage-français.xlsx'))
from openpyxl import load_workbook

def create_product_reference_comboboxes(frame, nombre_colis,parcel_data):
    product_comboboxes = []

    for i in range(int(nombre_colis)):
        product_references = []

        for j in range(int(parcel_data[i]["nombre_produits"].get())):
            # Create a ComboBox for each product reference
            label = ctk.CTkLabel(frame, text=f"Produit {j + 1} pour Colis {i + 1}:")
            label.grid(row=12 + j * 2, column=i, padx=5, pady=5)

            # Create a ComboBox with a list of product references
            product_combobox = ctk.CTkComboBox(frame, values=get_product_reference())
            product_combobox.grid(row=13+ j * 2, column=i, padx=5, pady=5)
            print(f"product_combobox {product_combobox}")
            # Add this ComboBox to the list
            product_references.append(product_combobox)
        print(f"product_comboboxes {product_references}")
        # Add the list of product references for this colis
        parcel_data[i]["product_references"] = product_references
        print(f"parcel_data {parcel_data}")
    
    bouton_soumettre = ctk.CTkButton(frame, text="Soumettre", command=lambda:  calculate_total_weight(frame, parcel_data, nombre_colis))
    bouton_soumettre.grid()
    
def calculate_total_weight(frame, parcel_data, nombre_colis):
    database = db.Database()
    list_poids_total = []

    for i in range(int(nombre_colis)):
        product_references = parcel_data[i]["product_references"] # Get product references for this colis
        print(f"product_references {product_references}")
        # Calculate total weight using product references for this colis
        total_weight_by_reference = calculate_total_weight_by_reference(database, product_references)
       
        list_poids_total.append(total_weight_by_reference)
        print(f"list_poids_total {list_poids_total}")

        print(f"Total Weight by Reference for Colis {i + 1}: {total_weight_by_reference}")
    for i in range(len(list_poids_total)):
        parcel_data[i]["Poids_total"]=list_poids_total[i]
        
    print(f"parcel_data {parcel_data}")
     
    # Update the Excel file with the total weights
    update_excel_file(nombre_colis, parcel_data,list_poids_total)

    # Optionally, display the total weights to the user
    display_total_weights(frame, nombre_colis, list_poids_total)

def calculate_total_weight_by_reference(database, product_references):
    total_weight = 0
    for reference in product_references:
        poids_produit = database.get_poids(reference.get())[0]
        print(f"poids_produit {poids_produit}")
        if poids_produit:
            total_weight += int(poids_produit)
            print(f"total_weight {total_weight}")
    return total_weight




def update_excel_file(nombre_colis, parcel_data,list_poids_total):
    # Load the Excel file
    wb = load_workbook(resource_path('CONTROLLER\\Modèle-liste-de-colisage-français.xlsx'))
    ws = wb.active
    import datetime

    # Get the current date
    current_date = datetime.date.today()

    # Print the current date
    print(current_date)

    produit=" "
    
    for numero_colis in range(nombre_colis):
        print(f"numero_colis {numero_colis}")
        wb = load_workbook(resource_path('CONTROLLER\\Modèle-liste-de-colisage-français.xlsx'))
        ws = wb.active
        a=26+numero_colis
        print("a "+str(a))
        b="C"+str(a)
        print("b "+b)
        ws[b]=str(numero_colis)
       
        for i in range(len(parcel_data[numero_colis]["product_references"])):
            produit=produit+parcel_data[numero_colis]["product_references"][i].get()+","
        print("produit "+produit)
        d="D"+str(a)
        produit=parcel_data[numero_colis]["Description"].get()
        print(f"parcel_data[numero_colis][dimensions] {produit}")
        ws[d]=parcel_data[numero_colis]["Description"].get()
        e="E"+str(a)
       
        dimension=parcel_data[numero_colis]["dimensions"].get()
        print(f"parcel_data[numero_colis][dimensions] {dimension}")
        ws[e]=parcel_data[numero_colis]["dimensions"].get()
        

        f="F"+str(a)
        Poids_total=parcel_data[numero_colis]["Poids_total"]
        print(f"parcel_data[numero_colis][Poids_total] {Poids_total }")
        ws[f]=parcel_data[numero_colis]["Poids_total"]
       

        g="G"+str(a)
        marque_colis=parcel_data[numero_colis]["marque_colis"].get()
        print(f"parcel_data[numero_colis][marque_colis] {marque_colis}")
        ws[g]=parcel_data[numero_colis]["marque_colis"].get()
        

        ws['F4'] = current_date
        ws['F12'] = sum(list_poids_total)

        # Enregistrer le fichier Excel
        wb.save(resource_path('CCONTROLLER\\Modèle-liste-de-colisage-français.xlsx'))
   

    
def display_total_weights(frame, nombre_colis, list_poids_total):
    label = ctk.CTkButton(frame, text="Ouvrir le fichier excel",command=fichier_excel)
    label.grid(padx=5,pady=5)
    CTkMessagebox(title="Succes", message="Vous pouvez consulter le fichier excel de liste de colisage")
def fichier_excel():
    import os
    # Spécifiez le chemin vers votre fichier Excel
    file_path = "CONTROLLER\\Modèle-liste-de-colisage-français.xlsx" 
    # Utilisez os.startfile pour ouvrir le fichier avec le programme par défaut
    os.startfile(file_path)
def get_product_reference():
    database=db.Database()
    List_references=[]
    list_reference=database.get_product_references()
    for i in  range(len(list_reference)):
        List_references.append(list_reference[i][0]) 
    return List_references 


       
def insert_data(root):
    # Créer une fenêtre tkinter
    for widget in root.winfo_children():
         widget.destroy()
    root.title("Interface de Liste de Colisage")
    bg=Image.open(resource_path("images\\image1.png"))
    bg1 = ImageTk.PhotoImage(bg)
    fenetre = ctk.CTkCanvas(root, width=700, height=400)
    fenetre.create_image(0, 0, image=bg1, anchor='nw')
    fenetre.pack(fill="both",expand=True)

    

    frame=ctk.CTkScrollableFrame(master=fenetre,width=800,height=500,bg_color="blue",corner_radius=5,border_color="darkblue",border_width=5)
    frame.grid(row=0,column=7,padx=250,pady=50)



    etiquette_nom_expediteur = ctk.CTkLabel(frame, text="La liste de colisage",text_color="blue",font=("bold",20))
    etiquette_nom_expediteur.grid(row=0,column=1,padx=5,pady=5)
    
    # Créer des champs de texte et des étiquettes pour chaque champ
    etiquette_nom_expediteur = ctk.CTkLabel(frame, text="Nom de l'expéditeur:")
    etiquette_nom_expediteur.grid(row=1,column=0,padx=5,pady=5)
    champ_nom_expediteur = ctk.CTkEntry(frame)
    champ_nom_expediteur.grid(row=2,column=0,padx=5,pady=5)

    etiquette_nom_entreprise = ctk.CTkLabel(frame, text="Nom de l'entreprise:")
    etiquette_nom_entreprise.grid(row=3,column=0,padx=5,pady=5)
    champ_nom_entreprise = ctk.CTkEntry(frame)
    champ_nom_entreprise.grid(row=4,column=0,padx=5,pady=5)

    etiquette_adresse_entre = ctk.CTkLabel(frame, text="Adress de l'entreprise:")
    etiquette_adresse_entre.grid(row=5,column=0,padx=5,pady=5)
    champ_adress_entre = ctk.CTkEntry(frame)
    champ_adress_entre.grid(row=6,column=0,padx=5,pady=5)

    etiquette_Tel_entre = ctk.CTkLabel(frame, text="Tel de l'entreprise:")
    etiquette_Tel_entre.grid(row=7,column=0,padx=5,pady=5)
    champ_Tel_entre = ctk.CTkEntry(frame)
    champ_Tel_entre.grid(row=8,column=0,padx=5,pady=5)
    etiquette_email_entre = ctk.CTkLabel(frame, text="E_mail d'entreprise:")
    etiquette_email_entre.grid(row=9,column=0,)
    champ_email_entre= ctk.CTkEntry(frame)
    champ_email_entre.grid(row=10,column=0,padx=5,pady=5)



    etiquette_nom_destinatiare = ctk.CTkLabel(frame, text="Nom de destinataire:")
    etiquette_nom_destinatiare.grid(row=1,column=1,padx=5,pady=5)
    champ_nom_destinataire = ctk.CTkEntry(frame)
    champ_nom_destinataire.grid(row=2,column=1,padx=5,pady=5)
    
    etiquette_Contact = ctk.CTkLabel(frame, text="Contact:")
    etiquette_Contact.grid(row=3,column=1,padx=5,pady=5)
    champ_Contact = ctk.CTkEntry(frame)
    champ_Contact.grid(row=4,column=1,padx=5,pady=5)

    etiquette_adress_destinataire = ctk.CTkLabel(frame, text="Adress de destinataire:")
    etiquette_adress_destinataire.grid(row=5,column=1,padx=5,pady=5)
    champ_adress_destinataire = ctk.CTkEntry(frame)
    champ_adress_destinataire.grid(row=6,column=1,padx=5,pady=5)

    etiquette_Tel_destinataire= ctk.CTkLabel(frame, text="Tel de destinataire:")
    etiquette_Tel_destinataire.grid(row=7,column=1,padx=5,pady=5)
    champ_Tel_destinataire = ctk.CTkEntry(frame)
    champ_Tel_destinataire.grid(row=8,column=1,padx=5,pady=5)

    etiquette_portabl_destinataire = ctk.CTkLabel(frame, text="portable de destinataire:")
    etiquette_portabl_destinataire.grid(row=9,column=1,padx=5,pady=5)
    champ_portabl_destinataire = ctk.CTkEntry(frame)
    champ_portabl_destinataire.grid(row=10,column=1,padx=5,pady=5)

    etiquette_email_destinataire = ctk.CTkLabel(frame, text="E_mail de destinataire:")
    etiquette_email_destinataire.grid(row=1,column=2,padx=5,pady=5)
    champ_email_destinataire = ctk.CTkEntry(frame)
    champ_email_destinataire.grid(row=2,column=2,padx=5,pady=5)

    etiquette_adress_destinataire = ctk.CTkLabel(frame, text="Adress de destinataire:") 
    etiquette_adress_destinataire.grid(row=3,column=2,padx=5,pady=5)
    champ_adress_destinataire = ctk.CTkEntry(frame)
    champ_adress_destinataire.grid(row=4,column=2,padx=5,pady=5)


    etiquette_numero_facture= ctk.CTkLabel(frame, text="Numero de facture:")
    etiquette_numero_facture.grid(row=5,column=2,padx=5,pady=5)
    champ_numero_facture = ctk.CTkEntry(frame)
    champ_numero_facture.grid(row=6,column=2,padx=5,pady=5)

    etiquette_numero_awb = ctk.CTkLabel(frame, text="Numero AWB:")
    etiquette_numero_awb.grid(row=7,column=2,padx=5,pady=5)
    champ_numero_awb = ctk.CTkEntry(frame)
    champ_numero_awb.grid(row=8,column=2,padx=5,pady=5)

    """etiquette_poids_brut = ctk.CTkLabel(frame, text="Poids brut:")
    etiquette_poids_brut.grid(row=11,column=2,padx=5,pady=5)
    champ_poids_brut = ctk.CTkEntry(frame)
    champ_poids_brut.grid(row=12,column=2,padx=5,pady=5)"""

    etiquette_nombre_colis = ctk.CTkLabel(frame, text="Nombre total de colis:")
    etiquette_nombre_colis.grid(row=9,column=2,padx=5,pady=5)
    champ_nombre_colis = ctk.CTkEntry(frame)
    champ_nombre_colis.grid(row=10,column=2,padx=5,pady=5)

    etiquette_raison= ctk.CTkLabel(frame, text="Raison de l'export:")
    etiquette_raison.grid(row=1,column=3,padx=5,pady=5)
    champ_raison = ctk.CTkEntry(frame)
    champ_raison.grid(row=2,column=3,padx=5,pady=5)

    etiquette_details = ctk.CTkLabel(frame, text="Autres details:")
    etiquette_details.grid(row=3,column=3,padx=5,pady=5)
    champ_details = ctk.CTkEntry(frame)
    champ_details.grid(row=4,column=3,padx=5,pady=5)

    # Créer un bouton pour enregistrer les données
    bouton_enregistrer = ctk.CTkButton(frame, text="Enregistrer", command=lambda:enregistrer(frame,champ_nom_expediteur.get(),champ_nom_entreprise.get(),champ_adress_entre.get(),champ_Tel_entre.get(),champ_email_entre.get(),champ_nom_destinataire.get(),champ_Contact.get(),champ_adress_destinataire.get(),champ_Tel_destinataire.get(),champ_portabl_destinataire.get(),champ_email_destinataire.get(),champ_numero_facture.get(),champ_numero_awb.get(),champ_nombre_colis.get(),champ_raison.get(),champ_details.get(),fenetre,root))
    bouton_enregistrer.grid(row=5,column=3,padx=5,pady=5)
    bouton_enregistrer = ctk.CTkButton(frame, text="Retourner", command=lambda:cl.retourner_Login(root))
    bouton_enregistrer.grid(row=5,column=4,padx=5,pady=5)

    # Lancer l'interface utilisateur
    root.mainloop()


#insert_data()




"""def enregistrer(frame, champ_nom_expediteur, champ_nom_entreprise, champ_adress_entre, champ_Tel_entre,
               champ_email_entre, champ_nom_destinataire, champ_Contact, champ_adress_destinataire,
               champ_Tel_destinataire, champ_portabl_destinataire, champ_email_destinataire,
               champ_numero_facture, champ_numero_awb, champ_nombre_colis, champ_raison, champ_details):
    nombre_colis = champ_nombre_colis
    if not nombre_colis.isdigit():
        label = ctk.CTkLabel(master=frame, text="Veuillez entrer un nombre", text_color="red", font=('bold', 10))
        label.grid(row=15, column=2)
    else:
        nombre_colis = int(nombre_colis)
        for widget in frame.winfo_children():
            widget.destroy()
        entry_list = []

        # For each colis
        for i in range(int(nombre_colis)):
            # Nombre de produits
            etiquette_produits = ctk.CTkLabel(frame, text=f"Nombre de produits pour colis {i}:")
            etiquette_produits.grid()
            
            # Create a StringVar for this Entry widget
            produits_var = StringVar()
            champ_produits = ctk.CTkEntry(frame, textvariable=produits_var)
            champ_produits.grid()

            # Dimensions en cm de colis
            etiquette_dimension = ctk.CTkLabel(frame, text=f"Dimensions en cm de colis {i}:")
            etiquette_dimension.grid()
            
            # Create a StringVar for this Entry widget
            dimension_var = StringVar()
            champ_dimension = ctk.CTkEntry(frame, placeholder_text="ex: 40x35X85", textvariable=dimension_var)
            champ_dimension.grid()

            # Marques sur colis
            etiquette_marque_colis = ctk.CTkLabel(frame, text=f"Marque sur colis {i}:")
            etiquette_marque_colis.grid()
            
            # Create a StringVar for this Entry widget
            marque_colis_var = StringVar()
            champ_marque_colis = ctk.CTkEntry(frame, placeholder_text="ex: 01/03", textvariable=marque_colis_var)
            champ_marque_colis.grid()

            # Append the StringVar variables to entry_list
            entry_list.append((produits_var, dimension_var, marque_colis_var))

        # Create a button to submit
        bouton_soumettre = ctk.CTkButton(frame, text="Soumettre",
                                        command=lambda: selectionner_produits(frame, nombre_colis, entry_list,
                                                                              bouton_soumettre))
        bouton_soumettre.grid()

# ..."""

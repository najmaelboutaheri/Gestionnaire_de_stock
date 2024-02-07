import tkinter as tk
from tkinter import ttk
import openpyxl
from CTkMessagebox import CTkMessagebox
import sys
import os
import getpass

# Obtenez le nom d'utilisateur actuel
current_user = getpass.getuser()

from MODEL import DatabaseSetup as db
from customtkinter import filedialog
from CONTROLLER import cLogin as cl
import sqlite3
themes_loaded = False
create_root=False

def load_themes(root):
    global themes_loaded
    if not themes_loaded:
        script_directory = os.path.dirname(os.path.abspath(__file__))
        light_theme_path = os.path.join(script_directory, "forest-light.tcl")
        dark_theme_path = os.path.join(script_directory, "forest-dark.tcl")

        root.tk.call("source", light_theme_path)
        root.tk.call("source", dark_theme_path)

        themes_loaded = True

def load_excel(treeview,mypath,mysheet):
    path =mypath 
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[mysheet]
    list_values = list(sheet.values)
    print(list_values)
    database=db.Database()

    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        #insert excel data into database
        Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille,nom_classification= value_tuple
        #verify if the sub_category name and cat_name exist in the database 
        if verification_cat_name(cat_name)==False:
           cl.historique(f'L\'utilisateur a ajoute une sous categorie  {cat_name}')
           database.insert_category(cat_name,nom_classification)
        if verification_subcategory_name(sub_cat_name)==False:
           cl.historique(f'L\'utilisateur a ajoute une sous categorie  {sub_cat_name}')
           database.insert_sub_category(sub_cat_name,cat_name)
        if get_product(Reference)==[]:
           cl.historique(f'L\'utilisateur a insere un produit de reference {Reference}')
           database.insert_product(Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille,nom_classification)
           #insert data into treeview
           treeview.insert('', tk.END, values=value_tuple)
        else:
           #insert data into treeview
           treeview.insert('', tk.END, values=value_tuple)


def load_data(treeview,sub_cat):
    lists_values=("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    for col_name in lists_values:
        treeview.heading(col_name, text=col_name)
    list_values=get_product_data(sub_cat)
    print(list_values)
    for value_tuple in list_values[0:]:
        treeview.insert('', tk.END, values=value_tuple)    

def insert_row(entry_values, treeview,entry_value):
    # Insert data into the database
    database=db.Database()
    Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille,nom_classification= entry_values
    if verification_cat_name(cat_name)==False:
        cl.historique(f'L\'utilisateur a ajoute une categorie {cat_name}')
        database.insert_category(cat_name,nom_classification)
    if verification_subcategory_name(sub_cat_name)==False:
        cl.historique(f'L\'utilisateur a ajoute une sous categorie  {sub_cat_name}')
        database.insert_sub_category(sub_cat_name,cat_name)
    if get_product(Reference)==[]:
        cl.historique(f'L\'utilisateur a insere un produit de reference {Reference}')
        #insert data into treeview
        database.insert_product(Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille,nom_classification) 
        for entry in entry_value:
            entry.delete(0, "end")
        #Insert row into the Treeview
        treeview.insert('', tk.END, values=entry_values)
    else:
        for entry in entry_value:
            entry.delete(0, "end")
        #Insert row into the Treeview
        treeview.insert('', tk.END, values=entry_values)
    
def insert_row_excel(entry_values, treeview,entry_value,path,sheet):
    # Insert data into the database
    database=db.Database()
    Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille,nom_classification= entry_values
   
    workbook = openpyxl.load_workbook(path)
    sheet =workbook[sheet]    
    sheet.append(entry_values)
    try:
        workbook.save(path)
    except PermissionError as e:
        CTkMessagebox(title="Warning", message="Vous devez fermer le fichier excel selectionne pour pouvoir l'editer")
    if verification_cat_name(cat_name)==False:
        cl.historique(f'L\'utilisateur a ajoute une categorie  {cat_name}')
        database.insert_category(cat_name,nom_classification)
    if verification_subcategory_name(sub_cat_name)==False:
        cl.historique(f'L\'utilisateur a ajoute une sous categorie  {sub_cat_name}')
        database.insert_sub_category(sub_cat_name,cat_name)
    if get_product(Reference)==[]:
        cl.historique(f'L\'utilisateur a insere un produit de reference {Reference}')
        #insert data into treeview
        database.insert_product(Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille,nom_classification) 
        for entry in entry_value:
            entry.delete(0, "end")
        #Insert row into the Treeview
        treeview.insert('', tk.END, values=entry_values)
    else:
        for entry in entry_value:
            entry.delete(0, "end")
        #Insert row into the Treeview
        treeview.insert('', tk.END, values=entry_values)

def Delete_row_excel(frame, widgets_frame, treeview, treeFrame, treeScroll, style,path,sheet,sub_cat):#frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet
    # Clear any existing widgets in the widgets_frame
    for widget in widgets_frame.winfo_children():
        widget.destroy()

    entry_label = ttk.Label(widgets_frame, text="Reference")
    entry_label.grid(sticky="w", padx=5, pady=5)
    entry = ttk.Entry(widgets_frame)
    entry.grid(sticky="w", padx=5, pady=5)

    def delete_and_refresh():
        Reference = entry.get()
        entry.delete(0,"end")
        database = db.Database()
        try:
            suppression = database.delete_product(Reference)
            # Your database operation that might raise an IntegrityError
            suppression = database.delete_product(Reference)
            if suppression:
                cl.historique(f'L\'utilisateur a supprime un produit de reference {Reference}')
                delete_row_by_reference(path, sheet, Reference) 
                # Clear existing data in the treeview
                for item in treeview.get_children():
                    treeview.delete(item)
                # Refresh the treeview content
                load_excel(treeview,path,sheet)
            else:
               CTkMessagebox(title="Warning", message="L'opération a échoué")
        except sqlite3.IntegrityError as e:
             CTkMessagebox(title="Warning", message="Impossible de supprimer ce produit")

        
    delete_button = ttk.Button(widgets_frame, text="Supprimer un produit",command=delete_and_refresh)
    delete_button.grid(row=3, column=2, padx=20, pady=10)

    back_button = ttk.Button(widgets_frame, text="Retourner", command=lambda: retourner_excel(path, sheet, widgets_frame, treeview, treeFrame, treeScroll, frame, style,sub_cat))
    back_button.grid(row=4, column=2, padx=20, pady=10)

def Delete_row(frame, widgets_frame, treeview, treeFrame, treeScroll, style,sub_cat):
    # Clear any existing widgets in the widgets_frame
    for widget in widgets_frame.winfo_children():
        widget.destroy()

    entry_label = ttk.Label(widgets_frame, text="Reference")
    entry_label.grid(sticky="w", padx=5, pady=5)
    entry = ttk.Entry(widgets_frame)
    entry.grid(sticky="w", padx=5, pady=5)

    def delete(sub_cat):
        Reference = entry.get()
        entry.delete(0,"end")
        database = db.Database()

        try:
        # Your database operation that might raise an IntegrityError
            suppression = database.delete_product(Reference)
            if suppression:
               cl.historique(f'L\'utilisateur a insere un produit de reference {Reference}')
               # Clear existing data in the treeview
               for item in treeview.get_children():
                   treeview.delete(item)
               # Refresh the treeview content
               load_data(treeview,sub_cat)
            else:
               CTkMessagebox(title="Warning", message="L'opération a échoué")
        except sqlite3.IntegrityError as e:
             CTkMessagebox(title="Warning", message="Impossible de supprimer ce produit")

    
    delete_button = ttk.Button(widgets_frame, text="Supprimer un produit",command=lambda:delete(sub_cat))
    delete_button.grid(row=3, column=2, padx=20, pady=10)

    back_button = ttk.Button(widgets_frame, text="Retourner", command=lambda: retourner( widgets_frame, treeview, treeFrame, treeScroll, frame, style,sub_cat))
    back_button.grid(row=4, column=2, padx=20, pady=10)

def delete_row_by_reference(file_path, sheet_name, reference_to_delete):
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        # Find the row with the reference and delete it
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == reference_to_delete:
                sheet.delete_rows(row[0].row, 1)  # Delete the entire row
        # Save the changes
        try:
            workbook.save(file_path)
        except PermissionError as e:
            CTkMessagebox(title="Warning", message="Vous devez fermer le fichier excel selectionne pour pouvoir l'editer")
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False
   
def search_row(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat):
    widgets_frame.destroy()
    widgets_frame = ttk.LabelFrame(frame, text="Vous pouvez rechercher un produit")
    widgets_frame.grid(row=0, column=2, padx=20, pady=10)
    entry_label = ttk.Label(widgets_frame, text="Reference")
    entry_label.grid(sticky="w", padx=5, pady=5)
    entry = ttk.Entry(widgets_frame)
    entry.grid(sticky="w", padx=5, pady=5)
    
    delete_button = ttk.Button(widgets_frame, text="rechercher un produit",command=lambda: search_Row(entry.get(),treeview,treeFrame,treeScroll,entry))#frame,widgets_frame ,treeview,path,mysheet
    delete_button.grid(row=3, column=2, padx=20, pady=10)
    back_button = ttk.Button(widgets_frame, text="Retourner",command=lambda:retourner(widgets_frame,treeview,treeFrame,treeScroll,frame,style,sub_cat))#frame,widgets_frame ,treeview,path,mysheet
    back_button.grid(row=4, column=2, padx=20, pady=10)

def search_row_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet,sub_cat):
    widgets_frame.destroy()
    widgets_frame = ttk.LabelFrame(frame, text="Vous pouvez rechercher un produit")
    widgets_frame.grid(row=0, column=2, padx=20, pady=10)
    entry_label = ttk.Label(widgets_frame, text="Reference")
    entry_label.grid(sticky="w", padx=5, pady=5)
    entry = ttk.Entry(widgets_frame)
    entry.grid(sticky="w", padx=5, pady=5)
    
    delete_button = ttk.Button(widgets_frame, text="rechercher un produit",command=lambda: search_Row_excel(entry.get(),treeview,treeFrame,treeScroll,entry,path,sheet))
    delete_button.grid(row=3, column=2, padx=20, pady=10)
    back_button = ttk.Button(widgets_frame, text="Retourner",command=lambda: retourner_excel(path, sheet, widgets_frame, treeview, treeFrame, treeScroll, frame, style,sub_cat))
    back_button.grid(row=4, column=2, padx=20, pady=10)


def search_Row(Reference, treeview, treeFrame, treeScroll, entry):
    # Connect to the database
    database = db.Database()
    
    # Get product details from the database
    Recherche = database.get_product_details(Reference)
    entry.delete(0, "end")
    
    if Recherche!=[]:
        cl.historique(f'L\'utilisateur a cherche sur  un produit de reference {Reference}')
        # Clear the existing data in the treeview
        for item in treeview.get_children():
            treeview.delete(item)
        
        # Update the treeview with matching rows
        for row in Recherche:
            treeview.insert('', tk.END, values=row)
    else:
        CTkMessagebox(title="Warning", message="Aucun produit trouvé")

def search_Row_excel(Reference, treeview, treeFrame, treeScroll, entry,path,sheet):
    # Connect to the database
    database = db.Database()
    
    # Get product details from the database
    Recherche = database.get_product_details(Reference)
    entry.delete(0, "end")
    
    if Recherche!=[]:
        # Search for a product in the Excel file based on the keyword (Reference)
        matching_rows = search_product_by_keyword(path, sheet, Reference)
        
        # Clear the existing data in the treeview
        for item in treeview.get_children():
            treeview.delete(item)
        
        # Update the treeview with matching rows
        for row in matching_rows:
            treeview.insert('', tk.END, values=row)
    else:
        CTkMessagebox(title="Warning", message="Aucun produit trouvé")

    

def search_product_by_keyword(file_path, sheet_name, keyword):
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Initialize a list to store matching rows
        matching_rows = []

        # Iterate through the rows and check if the keyword is present in any cell
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(keyword.lower() in str(cell).lower() for cell in row):
                matching_rows.append(row)

        # Return the list of matching rows
        return matching_rows

    except Exception as e:
        print(f"Error: {e}")
        return []


def clear_frame(frame):
    # Destroy all widgets in the frame
    for widget in frame.winfo_children():
        widget.destroy()
     

def toggle_mode(mode_switch,style):
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")
        
def retourner_excel(path,sheet,widgets_frame,treeview,treeFrame,treeScroll,frame,style,sub_cat):
    for widget in widgets_frame.winfo_children():
        widget.destroy()
    treeview.destroy()
    widgets_frame.destroy()
    cols = ("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    treeview = ttk.Treeview(treeFrame, show="headings",yscrollcommand=treeScroll.set, columns=cols, height=13)

    # Set column widths
    for col in cols:
        treeview.column(col, width=100)
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_excel(treeview,path,sheet)
    widgets_frame = ttk.LabelFrame(frame, text="Vous pouvez ajouter, supprimer ou inserer un produit")
    widgets_frame.grid(row=0, column=2, padx=0, pady=0)
    cols = ("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    entry_values = []
    for col in cols:
        entry_label = ttk.Label(widgets_frame, text=col)
        entry_label.grid(sticky="w", padx=5, pady=5)

        entry = ttk.Entry(widgets_frame)
        entry.grid(sticky="w", padx=5, pady=1)
        entry_values.append(entry)

    insert_button = ttk.Button(widgets_frame, text="Inserer", command=lambda: insert_row_excel([e.get() for e in entry_values], treeview,entry_values,path,sheet))
    insert_button.grid(sticky="w", padx=5, pady=5)
    delete_button = ttk.Button(widgets_frame, text="Supprimer", command=lambda: Delete_row_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet,sub_cat))#frame,widgets_frame ,treeview,path,mysheet
    delete_button.grid(row=1, column=3, padx=20, pady=10)
    search_button = ttk.Button(widgets_frame, text="Rechercher", command=lambda: search_row_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet,sub_cat))
    search_button.grid(row=2, column=3, padx=20, pady=10)
    back_button = ttk.Button(widgets_frame, text="Retourner",command=lambda:retourner(widgets_frame,treeview,treeFrame,treeScroll,frame,style,sub_cat))
    back_button.grid(row=3, column=3, padx=20, pady=10)
def retourner(widgets_frame,treeview,treeFrame,treeScroll,frame,style,sub_cat):
    for widget in widgets_frame.winfo_children():
        widget.destroy()
    treeview.destroy()
    widgets_frame.destroy()
    cols = ("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    treeview = ttk.Treeview(treeFrame, show="headings",yscrollcommand=treeScroll.set, columns=cols, height=13)

    # Set column widths
    for col in cols:
        treeview.column(col, width=100)
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data(treeview,sub_cat)
    widgets_frame = ttk.LabelFrame(frame, text="Vous pouvez ajouter, supprimer ou inserer un produit")
    widgets_frame.grid(row=0, column=2, padx=0, pady=0)
    cols = ("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    entry_values = []
    for col in cols:
        entry_label = ttk.Label(widgets_frame, text=col)
        entry_label.grid(sticky="w", padx=5, pady=5)

        entry = ttk.Entry(widgets_frame)
        entry.grid(sticky="w", padx=5, pady=1)
        entry_values.append(entry)

    insert_button = ttk.Button(widgets_frame, text="Inserer", command=lambda: insert_row([e.get() for e in entry_values], treeview,entry_values))
    insert_button.grid(sticky="w", padx=5, pady=5)
    delete_button = ttk.Button(widgets_frame, text="Supprimer", command=lambda: Delete_row(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat))#frame,widgets_frame ,treeview,path,mysheet
    delete_button.grid(row=1, column=3, padx=20, pady=10)
    search_button = ttk.Button(widgets_frame, text="Rechercher", command=lambda: search_row(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat))
    search_button.grid(row=2, column=3, padx=20, pady=10)
    search_button = ttk.Button(widgets_frame, text="Inserer un fichier", command=lambda: inserer_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat))
    search_button.grid(row=3, column=3, padx=20, pady=10)
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
def inserer_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat):
    path = resource_path(filedialog.askopenfilename(filetypes=[("Fichiers Excel", "*.xlsx *.xls")]))
    insert_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sub_cat)
def insert_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sub_cat):
    for widget in widgets_frame.winfo_children():
        widget.destroy()
    widgets_frame.destroy()
    
    widgets_frame = ttk.LabelFrame(frame, text="Le nom de la fenetre de fichier excel")
    widgets_frame.grid(row=0, column=2, padx=20, pady=10)
    entry_label = ttk.Label(widgets_frame, text="Nom de fenetre")
    entry_label.grid(sticky="w", padx=5, pady=5)
    sheet_name = ttk.Entry(widgets_frame)
    sheet_name.grid(sticky="w", padx=5, pady=5)
    
    delete_button = ttk.Button(widgets_frame, text="Entrer",command=lambda:excel_data(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet_name.get(),sub_cat)) 
    delete_button.grid(row=3, column=2, padx=20, pady=10)
    back_button = ttk.Button(widgets_frame, text="Retourner",command=lambda:retourner(widgets_frame,treeview,treeFrame,treeScroll,frame,style,sub_cat))
    back_button.grid(row=4, column=2, padx=20, pady=10)

    
def excel_data(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet,sub_cat):
    for widget in widgets_frame.winfo_children():
        widget.destroy()
    treeview.destroy()
    widgets_frame.destroy()
    cols = ("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    treeview = ttk.Treeview(treeFrame, show="headings",yscrollcommand=treeScroll.set, columns=cols, height=13)
    # Set column widths
    for col in cols:
        treeview.column(col, width=100)
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_excel(treeview,path,sheet)
    
    widgets_frame = ttk.LabelFrame(frame, text="Vous pouvez ajouter, supprimer ou inserer un produit")
    widgets_frame.grid(row=0, column=2, padx=0, pady=0)

    entry_values = []
    for col in cols:
        entry_label = ttk.Label(widgets_frame, text=col)
        entry_label.grid(sticky="w", padx=5, pady=5)

        entry = ttk.Entry(widgets_frame)
        entry.grid(sticky="w", padx=5, pady=1)
        entry_values.append(entry)

    insert_button = ttk.Button(widgets_frame, text="Inserer", command=lambda: insert_row_excel([e.get() for e in entry_values], treeview,entry_values,path,sheet))
    insert_button.grid(sticky="w", padx=5, pady=5)
    delete_button = ttk.Button(widgets_frame, text="Supprimer", command=lambda: Delete_row_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet,sub_cat))#frame,widgets_frame ,treeview,path,mysheet
    delete_button.grid(row=1, column=3, padx=20, pady=10)
    search_button = ttk.Button(widgets_frame, text="Rechercher", command=lambda: search_row_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,path,sheet,sub_cat))
    search_button.grid(row=2, column=3, padx=20, pady=10)
    back_button = ttk.Button(widgets_frame, text="Retourner",command=lambda:retourner(widgets_frame,treeview,treeFrame,treeScroll,frame,style,sub_cat))
    back_button.grid(row=3, column=3, padx=20, pady=10)

def product_data(sub_cat):
    print(sub_cat)
    root=tk.Tk()
    script_directory = os.path.dirname(os.path.abspath(__file__))
    light_theme_path = os.path.join(script_directory, "forest-light.tcl")
    dark_theme_path = os.path.join(script_directory, "forest-dark.tcl")

    root.tk.call("source", light_theme_path)
    root.tk.call("source", dark_theme_path)
    style = ttk.Style(root)    
    style.theme_use("forest-dark")

    

    frame = ttk.Frame(root,width=1000,height=700)
    frame.pack()

    
    
    """separator = ttk.Separator(widgets_frame)
    separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")"""

    """ mode_switch = ttk.Checkbutton(widgets_frame, text="Mode", style="Switch", command=lambda:toggle_mode(mode_switch,style))
    mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")"""

    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")
    #Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille
    cols = ("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    treeview = ttk.Treeview(treeFrame, show="headings",yscrollcommand=treeScroll.set, columns=cols, height=13)

    # Set column widths
    for col in cols:
        treeview.column(col, width=100)
   
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data(treeview,sub_cat)
    widgets_frame = ttk.LabelFrame(frame, text="Vous pouvez ajouter, supprimer ou inserer un produit")
    widgets_frame.grid(row=0, column=2, padx=0, pady=0)
        
    """separator = ttk.Separator(widgets_frame)
    separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")"""

    """mode_switch = ttk.Checkbutton(widgets_frame, text="Mode", style="Switch", command=lambda:toggle_mode(mode_switch,style))
    mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")"""
    entry_values = []
    for col in cols:
        entry_label = ttk.Label(widgets_frame, text=col)
        entry_label.grid(sticky="w", padx=5, pady=5)

        entry = ttk.Entry(widgets_frame)
        entry.grid(sticky="w", padx=5, pady=1)
        entry_values.append(entry)

    insert_button = ttk.Button(widgets_frame, text="Inserer", command=lambda: insert_row([e.get() for e in entry_values], treeview,entry_values))
    insert_button.grid(sticky="w", padx=5, pady=5)
    delete_button = ttk.Button(widgets_frame, text="Supprimer", command=lambda: Delete_row(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat))#frame,widgets_frame ,treeview,path,mysheet
    delete_button.grid(row=1, column=3, padx=20, pady=10)
    search_button = ttk.Button(widgets_frame, text="Rechercher", command=lambda: search_row(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat))
    search_button.grid(row=2, column=3, padx=20, pady=10)
    search_button = ttk.Button(widgets_frame, text="Inserer un fichier", command=lambda: inserer_excel(frame,widgets_frame,treeview,treeFrame,treeScroll,style,sub_cat))
    search_button.grid(row=3, column=3, padx=20, pady=10)
    root.mainloop()
def get_product_data(sub_cat):
    database=db.Database()
    products_data=database.get_products_details(sub_cat)
    return products_data

def get_product(Reference):
    database=db.Database()
    products_data=database.get_product_details(Reference)
    return products_data

def verification_cat_name(cat_name):
    # selectionner cat_name from database
    database=db.Database()
    verification=database.get_category_name(cat_name)
    if verification!=[]:
        return True
    else:
        return False
#verifier si une classification est deja existe dans la base de donnees
def verification_classification_name(classification_name):
    # selectionner classification_name from database
    database=db.Database()
    verification=database.get_classification_name(classification_name)
    if verification!=[]:
        return True
    else:
        return False
#verifier si une sous categorie est deja existe dans la base de donnees
def verification_subcategory_name(sub_cat_name):
    # selectionner sub_cat_name from database
    database=db.Database()
    verification=database.get_sub_categorie(sub_cat_name)
    if verification!=[]:
        return True
    else:
        return False

def show_products():
    root=tk.Tk()
    script_directory = os.path.dirname(os.path.abspath(__file__))
    light_theme_path = os.path.join(script_directory, "forest-light.tcl")
    dark_theme_path = os.path.join(script_directory, "forest-dark.tcl")

    root.tk.call("source", light_theme_path)
    root.tk.call("source", dark_theme_path)
    style = ttk.Style(root)    
    style.theme_use("forest-dark")
    frame = ttk.Frame(root,width=1000,height=700)
    frame.pack()
    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")
    #Reference, prod_name, description, prod_quantity_STOCK, sub_cat_name, cat_name, poids, taille
    cols = ("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    treeview = ttk.Treeview(treeFrame, show="headings",yscrollcommand=treeScroll.set, columns=cols, height=13)

    # Set column widths
    for col in cols:
        treeview.column(col, width=100)
   
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data_products(treeview)
    root.mainloop()
        
def products():
    database=db.Database()
    products=database.show_products()
    return products

def load_data_products(treeview):
    lists_values=("Reference", "nom_produit", "description", "prod_quantity", "sub_category", "cat_name", "poids", "la_taille","nom_classification")
    for col_name in lists_values:
        treeview.heading(col_name, text=col_name)
    list_values=products()
    print(list_values)
    for value_tuple in list_values[0:]:
        treeview.insert('', tk.END, values=value_tuple)